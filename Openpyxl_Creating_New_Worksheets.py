import openpyxl
from openpyxl import Workbook

def create_workbook_with_sheets(file_name: str) -> None:
    """
    Creates a new workbook with multiple worksheets and 
    saves it to the specified file.

    :param file_name: Name of the file to save the 
    workbook as.
    """
    # Create a new workbook
    wb: Workbook = Workbook()
    
    # Get the active worksheet and rename it
    ws1 = wb.active
    ws1.title = "MainSheet"

    # Create new worksheets
    ws2 = wb.create_sheet(title="Sheet2")
    ws3 = wb.create_sheet(title="Sheet3")

    # Add some data to each worksheet
    ws1['A1'] = "This is the Main Sheet"
    ws2['A1'] = "This is Sheet 2"
    ws3['A1'] = "This is Sheet 3"

    # Save the workbook
    wb.save(file_name)
    print(f"Workbook '{file_name}' with multiple sheets created" 
          "and saved successfully.")

# Specify the file name
file_name = "workbook_with_sheets.xlsx"

# Create the new workbook with multiple sheets
create_workbook_with_sheets(file_name)
