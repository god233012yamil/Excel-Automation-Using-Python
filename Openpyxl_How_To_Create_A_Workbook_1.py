import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

def create_new_workbook(file_name: str) -> None:
    """
    Creates a new workbook with sample data and saves it to the specified file.

    :param file_name: Name of the file to save the workbook as.
    """
    # Create a new workbook
    wb: Workbook = Workbook()
    
    # Get the active worksheet
    ws = wb.active
    ws.title = "Sample Data"

    # Add data to the worksheet
    headers = ["Name", "Age", "Department"]
    data = [
        ["John Doe", 30, "Sales"],
        ["Jane Smith", 25, "Marketing"],
        ["Emily Davis", 22, "Development"]
    ]
    
    # Append headers
    ws.append(headers)
    
    # Append data rows
    for row in data:
        ws.append(row)

    # Apply styling to the headers
    header_font = Font(size=12, bold=True)
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    
    for cell in ws[1]:  # First row is the header row
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook
    wb.save(file_name)
    print(f"Workbook '{file_name}' created and saved successfully.")

# Specify the file name
file_name = "new_workbook.xlsx"

# Create the new workbook
create_new_workbook(file_name)
