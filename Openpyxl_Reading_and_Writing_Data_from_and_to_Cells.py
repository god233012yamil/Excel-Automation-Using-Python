import openpyxl
from openpyxl import Workbook

def create_and_modify_workbook(file_name: str) -> None:
    """
    Creates a new workbook, writes data to cells, reads data 
    from cells, and saves the workbook.

    :param file_name: Name of the file to save the workbook as.
    """
    # Create a new workbook
    wb: Workbook = Workbook()
    
    # Get the active worksheet
    ws = wb.active
    ws.title = "Sample Data"

    # Write data to cells
    ws['A1'] = "Name"
    ws['B1'] = "Age"
    ws['C1'] = "Department"
    
    ws['A2'] = "John Doe"
    ws['B2'] = 30
    ws['C2'] = "Sales"
    
    ws['A3'] = "Jane Smith"
    ws['B3'] = 25
    ws['C3'] = "Marketing"

    # Save the workbook
    wb.save(file_name)
    print(f"Workbook '{file_name}' created and data written" 
          "successfully.")

    # Load the workbook
    wb = openpyxl.load_workbook(file_name)
    
    # Get the active worksheet
    ws = wb.active
    
    # Read data from cells
    name1 = ws['A2'].value
    age1 = ws['B2'].value
    department1 = ws['C2'].value

    name2 = ws['A3'].value
    age2 = ws['B3'].value
    department2 = ws['C3'].value

    # Print the read data
    print(f"Row 2: {name1}, {age1}, {department1}")
    print(f"Row 3: {name2}, {age2}, {department2}")

# Specify the file name
file_name = "workbook_read_write.xlsx"

# Create and modify the workbook
create_and_modify_workbook(file_name)
