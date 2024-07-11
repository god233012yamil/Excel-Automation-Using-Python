import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Protection

def create_workbook_with_validation_and_protection(file_name: str) -> None:
    """
    Creates a new workbook, adds data validation, protects the worksheet, 
    and saves the workbook.

    :param file_name: Name of the file to save the workbook as.
    """
    # Create a new workbook
    wb: Workbook = Workbook()
    
    # Get the active worksheet
    ws = wb.active
    ws.title = "Data Validation"

    # Add data to the worksheet
    ws['A1'] = "Name"
    ws['A2'] = "John Doe"
    ws['A3'] = "Jane Smith"
    ws['A4'] = "Emily Davis"

    ws['B1'] = "Age"
    ws['B2'] = 30
    ws['B3'] = 25
    ws['B4'] = 22

    ws['C1'] = "Department"
    ws['C2'] = "Sales"
    ws['C3'] = "Marketing"
    ws['C4'] = "Development"

    # Add data validation to the Age column to ensure values are between 18 and 65
    age_validation = DataValidation(type="whole", operator="between", formula1=18, 
                                    formula2=65, showErrorMessage=True)
    age_validation.error = 'Age must be between 18 and 65'
    ws.add_data_validation(age_validation)
    age_validation.add('B2:B4')

    # Protect the worksheet
    ws.protection.sheet = True
    ws.protection.password = 'mypassword'

    # Protect specific cells
    for cell in ws['A2:A4']:
        cell[0].protection = Protection(locked=False)
    
    # Save the workbook
    wb.save(file_name)
    print(f"Workbook '{file_name}' with data validation and protection created and" 
          "saved successfully.")

# Specify the file name
file_name = "workbook_with_validation_and_protection.xlsx"

# Create the new workbook with data validation and protection
create_workbook_with_validation_and_protection(file_name)
