import openpyxl
from openpyxl import Workbook

def create_workbook_with_formulas(file_name: str) -> None:
    """
    Creates a new workbook, adds data, includes formulas, and 
    saves the workbook.

    :param file_name: Name of the file to save the workbook as.
    """
    # Create a new workbook
    wb: Workbook = Workbook()
    
    # Get the active worksheet
    ws = wb.active
    ws.title = "Formulas and Functions"

    # Add data to the worksheet
    ws['A1'] = "Number 1"
    ws['A2'] = 10
    ws['A3'] = 20
    ws['A4'] = 30

    ws['B1'] = "Number 2"
    ws['B2'] = 40
    ws['B3'] = 50
    ws['B4'] = 60

    # Add a formula to calculate the sum of two numbers
    ws['C1'] = "Sum"
    ws['C2'] = "=A2+B2"
    ws['C3'] = "=A3+B3"
    ws['C4'] = "=A4+B4"

    # Add a formula to calculate the average
    ws['D1'] = "Average"
    ws['D2'] = "=AVERAGE(A2:A4)"
    ws['D3'] = "=AVERAGE(B2:B4)"

    # Save the workbook
    wb.save(file_name)
    print(f"Workbook '{file_name}' with formulas created and saved" 
          "successfully.")

# Specify the file name
file_name = "workbook_with_formulas.xlsx"

# Create the new workbook with formulas
create_workbook_with_formulas(file_name)
