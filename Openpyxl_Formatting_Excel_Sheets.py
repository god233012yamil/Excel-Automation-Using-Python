import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

def create_and_format_workbook(file_name: str) -> None:
    """
    Creates a new workbook, applies formatting to cells, 
    and saves the workbook.

    :param file_name: Name of the file to save the workbook as.
    """
    # Create a new workbook
    wb: Workbook = Workbook()
    
    # Get the active worksheet
    ws = wb.active
    ws.title = "Formatted Data"

    # Add data to the worksheet
    headers = ["Name", "Age", "Department"]
    data = [
        ["John Doe", 30, "Sales"],
        ["Jane Smith", 25, "Marketing"],
        ["Emily Davis", 22, "Development"]
    ]
    
    # Append headers
    ws.append(headers)
    
    # Append data rows
    for row in data:
        ws.append(row)

    # Apply formatting to the headers
    header_font = Font(size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", 
                              end_color="4F81BD", 
                              fill_type="solid")
    thin_border = Border(left=Side(style="thin"),
                         right=Side(style="thin"),
                         top=Side(style="thin"),
                         bottom=Side(style="thin"))
    
    for cell in ws[1]:  # First row is the header row
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook
    wb.save(file_name)
    print(f"Workbook '{file_name}' created and formatted successfully.")

# Specify the file name
file_name = "formatted_workbook.xlsx"

# Create and format the workbook
create_and_format_workbook(file_name)
